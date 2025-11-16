"""
Admin controller for handling admin-related requests
"""
from sqlalchemy.orm import Session
from fastapi import HTTPException, status
from app.models.admin import Admin
from app.models.interview import Interview
from app.models.candidate import Candidate
from app.models.answer import Answer
from app.models.question import Question
from app.models.interview_link import InterviewLink
from app.schemas.admin import AdminSignupRequest, AdminLoginRequest, AdminLoginResponse
from app.schemas.interview import InterviewCreate, InterviewResponse, InterviewListResponse
from app.schemas.answer import DashboardResponse, ReportResponse, CandidateDetail, CandidateAnswerDetail
from app.services.auth_service import (
    verify_password,
    get_password_hash,
    create_access_token,
    get_admin_by_email
)
from app.services.interview_service import create_interview_with_link
from app.services.email_service import send_interview_invitation
from typing import List, Optional
from datetime import datetime
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def signup_admin(db: Session, signup_data: AdminSignupRequest) -> AdminLoginResponse:
    """
    Create a new admin account and return JWT token
    
    Args:
        db: Database session
        signup_data: Signup credentials (email and password)
    
    Returns:
        AdminLoginResponse with JWT token
    
    Raises:
        HTTPException: If email already exists
    """
    # Check if admin with this email already exists
    existing_admin = get_admin_by_email(db, signup_data.email)
    if existing_admin:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
    
    # Create new admin
    password_hash = get_password_hash(signup_data.password)
    new_admin = Admin(
        email=signup_data.email,
        password_hash=password_hash
    )
    
    db.add(new_admin)
    db.commit()
    db.refresh(new_admin)
    
    # Create JWT token
    token_data = {
        "sub": str(new_admin.id),
        "email": new_admin.email,
        "type": "admin"
    }
    access_token = create_access_token(data=token_data)
    
    return AdminLoginResponse(
        access_token=access_token,
        admin_id=str(new_admin.id),
        email=new_admin.email
    )


def login_admin(db: Session, login_data: AdminLoginRequest) -> AdminLoginResponse:
    """
    Authenticate admin and return JWT token
    
    Args:
        db: Database session
        login_data: Login credentials
    
    Returns:
        AdminLoginResponse with JWT token
    
    Raises:
        HTTPException: If credentials are invalid
    """
    admin = get_admin_by_email(db, login_data.email)
    
    if not admin or not verify_password(login_data.password, admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid email or password"
        )
    
    # Create JWT token
    token_data = {
        "sub": str(admin.id),
        "email": admin.email,
        "type": "admin"
    }
    access_token = create_access_token(data=token_data)
    
    return AdminLoginResponse(
        access_token=access_token,
        admin_id=str(admin.id),
        email=admin.email
    )


def create_interview(
    db: Session,
    interview_data: InterviewCreate,
    base_url: str = "http://localhost:5173"
) -> InterviewResponse:
    """
    Create a new interview with shareable link
    
    Args:
        db: Database session
        interview_data: Interview creation data
        base_url: Base URL for generating shareable links
    
    Returns:
        InterviewResponse with shareable link
    """
    interview, link_code = create_interview_with_link(  # No password returned
        db=db,
        title=interview_data.title,
        description=interview_data.description
    )
    
    # Generate shareable link URL
    shareable_link = f"{base_url}/interview/{link_code}"
    
    # Simulate sending email (logs to console)
    print(f"\n[EMAIL SIMULATION] Interview created:")
    print(f"  Link: {shareable_link}\n")
    
    return InterviewResponse(
        id=str(interview.id),
        title=interview.title,
        description=interview.description,
        created_at=interview.created_at,
        shareable_link=shareable_link,
        candidate_password=None  # No password
    )


def add_question_to_interview(
    db: Session,
    interview_id: str,
    question_text: str
) -> dict:
    """
    Add a question to an interview
    
    Args:
        db: Database session
        interview_id: Interview ID
        question_text: Question text
    
    Returns:
        Dictionary with question details
    
    Raises:
        HTTPException: If interview not found
    """
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    if not interview:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Interview not found"
        )
    
    question = Question(
        interview_id=interview_id,
        question_text=question_text
    )
    db.add(question)
    db.commit()
    db.refresh(question)
    
    return {
        "id": str(question.id),
        "interview_id": str(question.interview_id),
        "question_text": question.question_text,
        "created_at": question.created_at
    }


def get_all_interviews(db: Session) -> InterviewListResponse:
    """
    Get all interviews
    
    Args:
        db: Database session
    
    Returns:
        InterviewListResponse with list of interviews
    """
    interviews = db.query(Interview).all()
    
    interview_responses = []
    for interview in interviews:
        print(f"Interview ID: {interview.id}, Title: {interview.title}, Description: {interview.description}, Created: {interview.created_at}")
        # Get shareable link if exists
        shareable_link = None
        if interview.interview_link:
            shareable_link = f"http://localhost:5173/interview/{interview.interview_link.link_code}"
        
        interview_responses.append(InterviewResponse(
            id=str(interview.id),
            title=interview.title,
            description=interview.description,
            created_at=interview.created_at,
            shareable_link=shareable_link,
            
            
        ))
    
    return InterviewListResponse(
        interviews=interview_responses,
        total=len(interview_responses)
    )


def get_interview_dashboard(db: Session, interview_id: str) -> DashboardResponse:
    """
    Get dashboard data for a specific interview
    
    Args:
        db: Database session
        interview_id: Interview ID
    
    Returns:
        DashboardResponse with interview stats and candidate list
    
    Raises:
        HTTPException: If interview not found
    """
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    if not interview:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Interview not found"
        )
    
    # Get all candidates for this interview
    candidates = db.query(Candidate).filter(Candidate.interview_id == interview_id).all()
    
    candidate_details = []
    for candidate in candidates:
        # Get answers for this candidate
        answers = db.query(Answer).filter(Answer.candidate_id == candidate.id).all()
        
        # Calculate average score
        scores = [a.score for a in answers if a.score is not None]
        average_score = sum(scores) / len(scores) if scores else None
        
        candidate_details.append(CandidateDetail(
            candidate_id=str(candidate.id),
            name=candidate.name,
            email=candidate.email,
            created_at=candidate.created_at,
            total_answers=len(answers),
            average_score=average_score
        ))
    
    # Get total questions
    total_questions = db.query(Question).filter(Question.interview_id == interview_id).count()
    
    return DashboardResponse(
        interview_id=str(interview.id),
        interview_title=interview.title,
        total_candidates=len(candidates),
        total_questions=total_questions,
        candidates=candidate_details
    )


def get_interview_candidates(db: Session, interview_id: str) -> List[dict]:
    """
    Get all candidates for an interview
    
    Args:
        db: Database session
        interview_id: Interview ID
    
    Returns:
        List of candidate dictionaries
    
    Raises:
        HTTPException: If interview not found
    """
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    if not interview:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Interview not found"
        )
    
    candidates = db.query(Candidate).filter(Candidate.interview_id == interview_id).all()
    
    return [
        {
            "id": str(c.id),
            "name": c.name,
            "email": c.email,
            "created_at": c.created_at
        }
        for c in candidates
    ]


def get_candidate_report(db: Session, candidate_id: str) -> ReportResponse:
    """
    Get detailed report for a candidate
    
    Args:
        db: Database session
        candidate_id: Candidate ID
    
    Returns:
        ReportResponse with candidate answers and scores
    
    Raises:
        HTTPException: If candidate not found
    """
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Candidate not found"
        )
    
    # Get interview
    interview = candidate.interview
    
    # Get all answers for this candidate
    answers = db.query(Answer).filter(Answer.candidate_id == candidate_id).all()
    
    answer_details = []
    for answer in answers:
        question = db.query(Question).filter(Question.id == answer.question_id).first()
        answer_details.append(CandidateAnswerDetail(
            answer_id=str(answer.id),
            question_text=question.question_text if question else "Unknown question",
            transcript=answer.transcript,
            score=answer.score,
            feedback=answer.feedback,
            created_at=answer.created_at
        ))
    
    # Calculate average score
    scores = [a.score for a in answers if a.score is not None]
    average_score = sum(scores) / len(scores) if scores else None
    
    # Get total questions in interview
    total_questions = db.query(Question).filter(Question.interview_id == candidate.interview_id).count()
    
    return ReportResponse(
        candidate_id=str(candidate.id),
        candidate_name=candidate.name,
        candidate_email=candidate.email,
        interview_id=str(candidate.interview_id),
        interview_title=interview.title,
        answers=answer_details,
        average_score=average_score,
        total_questions=total_questions,
        completed_questions=len(answers)
    )


def generate_interview_excel(db: Session, interview_id: str) -> BytesIO:
    """
    Generate Excel report for an interview with all candidates sorted by score
    
    Args:
        db: Database session
        interview_id: Interview ID
    
    Returns:
        BytesIO object containing Excel data
    
    Raises:
        HTTPException: If interview not found
    """
    interview = db.query(Interview).filter(Interview.id == interview_id).first()
    if not interview:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Interview not found"
        )
    
    # Get all candidates for this interview
    candidates = db.query(Candidate).filter(Candidate.interview_id == interview_id).all()
    
    # Build candidate data with scores and feedback
    candidate_data = []
    for candidate in candidates:
        answers = db.query(Answer).filter(Answer.candidate_id == candidate.id).all()
        scores = [a.score for a in answers if a.score is not None]
        average_score = sum(scores) / len(scores) if scores else 0.0
        
        # Collect all feedback
        feedback_list = []
        for answer in answers:
            if answer.feedback:
                question = db.query(Question).filter(Question.id == answer.question_id).first()
                if question:
                    feedback_list.append(f"Q: {question.question_text}\nA: {answer.feedback}")
        
        candidate_data.append({
            'name': candidate.name,
            'email': candidate.email,
            'score': average_score,
            'feedback': '\n\n'.join(feedback_list) if feedback_list else 'No feedback available'
        })
    
    # Sort candidates by score (highest first)
    candidate_data.sort(key=lambda x: x['score'], reverse=True)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Interview Report"
    
    # Define styles
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Header row
    headers = ['Rank', 'Candidate Name', 'Email', 'Score', 'Feedback']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Data rows
    for idx, candidate in enumerate(candidate_data, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=candidate['name']).border = border
        ws.cell(row=row, column=3, value=candidate['email']).border = border
        ws.cell(row=row, column=4, value=f"{candidate['score']:.1f}/10").border = border
        ws.cell(row=row, column=5, value=candidate['feedback']).border = border
        ws.cell(row=row, column=5).alignment = wrap_align
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 60
    
    # Set row heights for feedback column
    for row in range(2, len(candidate_data) + 2):
        ws.row_dimensions[row].height = 60
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

